from tkinter import Tk, ttk
from tkinter import filedialog
from tkinter import * 
from tkinter.ttk import *
from tkinter import messagebox

import sys
from tkinter.scrolledtext import ScrolledText

from tkinter.messagebox import showerror

import openpyxl
from openpyxl import *
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Font, Color, colors, PatternFill, Border, Side, Alignment, Protection
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Border, Side
from queue import Empty
from openpyxl.utils import get_column_letter
import itertools
from operator import concat
from queue import Empty
from openpyxl.utils import rows_from_range
from copy import copy



# Define set_border function
def set_border(ws, cell_range):
    rows = ws[cell_range]
    side = Side(border_style='thin', color="FF000000")

    rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
            if pos_x == 0:
                border.left = side
            if pos_x == max_x:
                border.right = side
            if pos_y == 0:
                border.top = side
            if pos_y == max_y:
                border.bottom = side

            # set new border only if it's one of the edge cells
            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border


class PrintLogger(object):  # create file like object

    def __init__(self, textbox):  # pass reference to text widget
        self.textbox = textbox  # keep ref

    def write(self, text):
        self.textbox.configure(state="normal")  # make field editable
        self.textbox.insert("end", text)  # write text to textbox
        self.textbox.see("end")  # scroll to end
        self.textbox.configure(state="disabled")  # make field readonly

    def flush(self):  # needed for file like object
        pass


class App(Tk):
    def __init__(self):
        super().__init__()

        #button1 = ttk.Button(self, text='S??lectioner un fichier de commande', command=self.browse_files)
        button1 = ttk.Button(self, text='S??lectioner un fichier de commande', command=lambda: [self.browse_files(), self.redirect_logging()])
        button1.grid(row=0, column=0)

        button2 = ttk.Button(self, text='G??n??rer le fichier producteurs', command=self.gen_prod)
        button2.grid(row=4, column=0)

        button3 = ttk.Button(self, text='G??n??rer le fichier clients', command=self.gen_cust)
        button3.grid(row=4, column=1)

        button4 = ttk.Button(self, text='Afficher le fichier s??lectionn??', command=self.show)
        button4.grid(row=0, column=1)

        self.filepath_label = ttk.Label(self, text="Fichier actuellement s??lectionn??:")
        self.filepath_label.grid(row=2, columnspan=2)     

        self.filepath = ttk.Label(self, foreground='green')
        self.filepath.grid(row=3, columnspan=2)   

        self.log_widget = ScrolledText(self, height=25, width=120, font=("courrier", "10", "normal"))
        self.log_widget.grid(row=9, columnspan=2)   

        self.protocol("WM_DELETE_WINDOW", self.on_closing)



    def redirect_logging(self):
        logger = PrintLogger(self.log_widget)
        sys.stdout = logger
        sys.stderr = logger
    
    def on_closing(self):
        if messagebox.askyesno(title="Quitter ?", message="Voulez-vous vraiment quitter ?"):
            self.destroy()
        
    def browse_files(self):
        # use instance variable self.filename
        self.filename = filedialog.askopenfilename(initialdir="/",
                                                title="Select a File",
                                                filetypes=((".xls", "*.xls"),
                                                            (".xlsx", "*.xlsx")))
        self.filepath.config(text=self.filename)


    def callback(self):
        if askyesno('Titre 1', '??tes-vous s??r de vouloir faire ??a?'):
            showwarning('Titre 2', 'Tant pis...')
        else:
            showinfo('Titre 3', 'Vous avez peur!')
            showerror("Titre 4", "Aha")

    def show(self):
        print(self.filename)


################################
#### SPLIT PER PRODUCTOR #######
################################


    def gen_prod(self):
        print(self.filename)

        # create a workbook
        wb = Workbook()
        #ws = wb.active

        # Give the location of the file
        path = self.filename

        # To open the workbook
        # workbook object is created
        wb_obj = openpyxl.load_workbook(path, data_only=True)

        # Get workbook active sheet object
        # from the active attribute
        sheet_obj = wb_obj.active

        # Getting the value of maximum rows
        # and column
        row = sheet_obj.max_row
        column = sheet_obj.max_column

        print("Total Rows:", row)
        print("Total Columns:", column)

        print("\nValue of first row - FAMILY NAME")
        for i in range(1, column + 1): 
            cell_obj = sheet_obj.cell(row = 5, column = i) 
            if cell_obj.value:
                print(cell_obj.value, end = " ")
            

        # printing the value of last row
        print("\nValue of last column - PRODUCTORS LIST")

        # initialize prod_name and prod_pos_row arrays
        prod_name = []
        prod_pos_row =[]
        sheets_to_remove = []
        position = 0



        # loop through rows and detect productors and position in file
        for a in range(1, row + 1):

            producteur = sheet_obj.cell(row = a, column = 1).value
            if producteur:
                if '"**"' in str(producteur):
                    producteur=(producteur.lstrip('"**"'))
                    prod_name.append(producteur)
                    prod_pos_row.append(a)

        # List productors and positions in file
        #print(prod_name)
        #print(prod_pos_row)

        # Print headers / columns titles
        #print("intitul??;description;unite;prix unit??;quantite_total;montant_total")

        #One sheet for all productors
        wb.create_sheet("Commandes globales")
        wsglob = wb["Commandes globales"]

        #Define copy range function
        def copy_range(range_str, src, dst):

            for row in rows_from_range(range_str):
                for cell in row:
                    dst[cell].value = src[cell].value
                    dst[cell].border = src[cell].border

            return

        total_item_array = [0]

        # Loop through file to list ordered items per productor
        for prod in range (0,len(prod_name)-1):

            total = 0
            total_item = 0

        #    print("Producteur ",prod,"/",len(prod_name))
            print("*** Producteur: ", prod_name[prod]," ***")
            wb.create_sheet(prod_name[prod])
            if prod == 1: del wb['Sheet']



            ws = wb[prod_name[prod]]
            ws['A1'] = prod_name[prod]
            ws['B1'] = 'Description'
            ws['C1'] = 'Unit??'
            ws['D1'] = 'Prix unitaire'
            ws['E1'] = 'Quantit??'
            ws['F1'] = 'Total'

            total_item = 0    

            for item in range (prod_pos_row[prod]+1,(prod_pos_row[prod+1])):

                montant_total = sheet_obj.cell(row = item, column = sheet_obj.max_column - 1)
                #print(prod,item,montant_total,montant_total.value)

                if montant_total.value: 

                    intitul?? = sheet_obj.cell(row = item, column = 1).value
                    description = sheet_obj.cell(row = item, column = 2).value
                    unite = sheet_obj.cell(row = item, column = 3).value
                    prixun = sheet_obj.cell(row = item, column = 4).value
                    quantite_total = sheet_obj.cell(row = item, column = sheet_obj.max_column - 2).value
                    montant_total = round(float(montant_total.value),2)

                    print(intitul??,";",description,";",unite,";",prixun,";",quantite_total,";",montant_total)

                    insert_row=intitul??,description,unite,prixun,quantite_total,montant_total
                    ws.append(insert_row)

                    total = total + round(float(montant_total),2)

                    #print(prod_name[prod], total)

                    if montant_total != 0:
                        total_item = total_item + 1
                
                # Check at the last row of a given productor if the total is 0. If so, this productor will not be displayed in report                                                        
                if item == (prod_pos_row[prod+1]-1) and total == 0:
                    print("!!!!!!!!!!!!!!!!!!!!!!!!!",prod_name[prod],item,prod_pos_row[prod])
                    print("TOTAL: ", total)
                    sheets_to_remove.append(prod_name[prod])



            total_item_array.append(total_item)

            ligne_total="TOTAL " + prod_name[prod],"","","","",total
            ws.append(ligne_total)

            # Merge last row for total display
            current_row = ws.max_row
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)

            # Highlight last row / total
            for cell in ws[current_row]:
                cell.font = Font(color="0e1643", bold=True, size=14)
                cell.fill = PatternFill('solid', fgColor = 'cdc8b1')


            #Define printing area
            last = ws.calculate_dimension()
            print("Print area for this sheet: ",last)
            ws.print_area = last


            def set_border(ws, cell_range):
                thin = Side(border_style="thin", color="000000")
                for row in ws[cell_range]:
                    for cell in row:
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            set_border(ws, last)


            wsprops = ws.sheet_properties
            wsprops.tabColor = "1072BA"
            wsprops.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)
            #wsprops.pageSetUpPr.autoPageBreaks = True    
            wsprops.print_title_rows='1:1'
            #print(wsprops)

            #Set landscape orientation
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

            for cell in ws["1:1"]:
        #                print(cell)
                cell.font = Font(color="0e1643", bold=True, size=14)
                cell.fill = (PatternFill('solid', fgColor = 'cdc8b1'))

            # Copy current productor in global worksheet / still needs to be positioned after previous productor (iteration)
            #copy_range(last,ws,wsglob)

            if prod == 0:
                position = 0
            else:
                position = position + total_item_array[prod] + 5

            if total_item != 0:

                for row in ws.rows:
                    for cell in row:
                        new_cell = wsglob.cell(row=cell.row + position, column=cell.column,
                                value= cell.value)
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = copy(cell.number_format)
                            new_cell.protection = copy(cell.protection)
                            new_cell.alignment = copy(cell.alignment)

                wsprops = wsglob.sheet_properties
                wsprops.tabColor = "93c47d"
                wsprops.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)
                wsprops.outlinePr.applyStyles = True

            # Auto-sizing columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try: # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 1) * 1.5
                ws.column_dimensions[column].width = adjusted_width
                #print("Cell width adjusted for : ", cell)

        for col in wsglob.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 1) * 1.5
            wsglob.column_dimensions[column].width = adjusted_width
            #print("Cell width adjusted for : ", cell)

        print ("Feuilles pr??sentes dans le fichier: ", wb.sheetnames)
        print("Feuilles vides ?? supprimer: ", sheets_to_remove)
        print("Nombre de feuilles vides ?? supprimer: ", len(sheets_to_remove))

        for rem in range(0,len(sheets_to_remove)):
            print(rem, sheets_to_remove[rem])
            wb.remove(wb[sheets_to_remove[rem]])


        output_filename =(path.replace(".xls", "-par producteur.xls"))
        wb.save(output_filename)

        print ("Nombre de produits command??s par producteur")
        for prod in range (0,len(prod_name)-1):
            print(prod_name[prod]," : ",total_item_array[prod+1])

        print("G??n??ration du fichier client termin??e :", output_filename)
        showerror(title='Error', message=("G??n??ration du fichier PRODUCTEURS termin??e: " + output_filename))
    


################################
#### SPLIT PER CUSTOMER ########
################################


    def gen_cust(self):

        # create a workbook
        wb = Workbook()
        #ws = wb.active

        # Give the location of the file
        path = self.filename

        # To open the workbook
        # workbook object is created
        wb_obj = openpyxl.load_workbook(path, data_only=True)

        # Get workbook active sheet object
        # from the active attribute
        sheet_obj = wb_obj.active

        # Getting the value of maximum rows
        # and column
        #row = sheet_obj.max_row
        column = sheet_obj.max_column

        for row in range(1, sheet_obj.max_row +1):
            if 'BE33000441432246' in str(sheet_obj.cell(row, 1).value):
                last_row = row
                print("Last row is: ", last_row)

        print("Total Rows:", row)
        print("Total Columns:", column)


        # initialize family and family_pos_row arrays
        family = []
        family_pos_col = []

        print("\n - FAMILIES with orders")
        for i in range(1, column + 1): 
            cell_obj = sheet_obj.cell(row = 5, column = i)
            print(cell_obj, cell_obj.value)
            orders = sheet_obj.cell(row = last_row, column = i)
            print(orders, orders.value)
            if cell_obj.value and orders.value != 0:
                print("Cell object value >< 0: ",cell_obj.value)
                family.append(cell_obj.value)
                family_pos_col.append(i)    

        # printing the value of last row
        print("\nValue of last column - PRODUCTORS LIST")

        # initialize prod_name and prod_pos_row arrays
        prod_name = []
        prod_pos_row =[]
        sheets_to_remove = []
        position = 0
        total_item = 0

        wb.create_sheet("Commandes globales")
        wsglob = wb["Commandes globales"]

        total_item_array = [0]


        # loop through rows and detect productors and position in file
        for a in range(1, row + 1):

            producteur = sheet_obj.cell(row = a, column = 1).value
            if producteur:
                if '"**"' in str(producteur):
                    producteur=(producteur.lstrip('"**"'))
                    prod_name.append(producteur)
                    prod_pos_row.append(a)


        # Loop through file to list ordered items per family
        for fam in range (0,len(family)):
            total = 0
            total_item = 0

            print("***** Famille: ", family[fam]," *****")
            #print("*** Colonne nr: ", family_pos_col[fam]," ***")
            wb.create_sheet(family[fam])

            #Remove default unused excel sheet
            if fam == 1: del wb['Sheet']

            ws = wb[family[fam]]

            ws['A1'] = family[fam]
            ws['B1'] = 'DESCRIPTION'
            ws['C1'] = 'UNITE'
            ws['D1'] = 'PRIX UNITAIRE'
            ws['E1'] = 'QUANTITE'
            ws['F1'] = 'TOTAL'
                
            for prod in range (0,len(prod_name)-1):

                print("** Producteur: ",prod_name[prod]," **")
                
                #Debugging lines
                #print ("Nom du producteur: ",prod_name[prod])
                #print ("Premi??re ligne du producteur: ",prod_pos_row[prod]+1)
                #print ("Premi??re ligne du producteur suivant: ",prod_pos_row[prod+1])

                ws.append((prod_name[prod],""))
                total_item = total_item + 1


                current_row = ws.max_row

                #print("Ligne actuelle: ", prod_name[prod], current_row)

                # Highlight productor row
                for cell in ws[current_row]:
                    cell.font = Font(color="0e1643", bold=True, size=12)
                    cell.fill = PatternFill('solid', fgColor = 'd0e0e3')

                total_prod = 0

                for item in range (prod_pos_row[prod]+1,(prod_pos_row[prod+1])):

                    intitul?? = sheet_obj.cell(row = item, column = 1).value
                    description = sheet_obj.cell(row = item, column = 2).value
                    unite = sheet_obj.cell(row = item, column = 3).value
                    prixun = sheet_obj.cell(row = item, column = 4).value
                    quantite_total = sheet_obj.cell(row = item, column = family_pos_col[fam]).value

                    
                    if quantite_total and float(quantite_total) > 0:

                        montant_total = float(quantite_total) * float(prixun)
                        montant_total=round(montant_total,2)

                        #print("\nITEM Ligne: ",item)
                        #print("ITEM Colonne: ",family_pos_col[fam])
                        #print("ITEM Famille: ",family[fam])
                        #print("ITEM quantite_total: ",quantite_total)
                        #print("ITEM prix unitaire: ",prixun)
                        #print("montant_total: ",montant_total)

                        print(intitul??,";",description,";",unite,";",prixun,";",quantite_total,";",montant_total)

                        insert_row=intitul??,description,unite,prixun,quantite_total,montant_total
                        ws.append(insert_row)

                        total = total + montant_total
                        total_prod = total_prod + montant_total

                        if montant_total != 0:
                            total_item = total_item + 1

                #print("Nom du producteur: ",prod_name[prod], "Total prod: ", total_prod, "Current row: ", current_row, "Total item: ",total_item)
                if total_prod == 0:
                    #print("ligne supprim??e: ",current_row)
                    ws.delete_rows(current_row)
                    total_item = total_item - 1



            ligne_total="TOTAL",family[fam],"IBAN","BE33000441432246","",total
            ws.append(ligne_total)

            total_item_array.append(total_item)

            current_row = ws.max_row

            # Highlight last row / total
            for cell in ws[current_row]:
                cell.font = Font(color="0e1643", bold=True, size=14)
                cell.fill = PatternFill('solid', fgColor = 'cdc8b1')
            

            #Define printing area
            last = ws.calculate_dimension()
            #print("Print area for this sheet: ",last)
            ws.print_area = last


            def set_border(ws, cell_range):
                thin = Side(border_style="thin", color="000000")
                for row in ws[cell_range]:
                    for cell in row:
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            set_border(ws, last)


            wsprops = ws.sheet_properties
            wsprops.tabColor = "1072BA"
            wsprops.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)
            #wsprops.pageSetUpPr.autoPageBreaks = True    
            wsprops.print_title_rows='1:1'
            #print(wsprops)

            #Set landscape orientation
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

            for cell in ws["1:1"]:
        #                print(cell)
                cell.font = Font(color="0e1643", bold=True, size=14)
                cell.fill = (PatternFill('solid', fgColor = 'cdc8b1'))

            # Copy current productor in global worksheet / still needs to be positioned after previous productor (iteration)
            #copy_range(last,ws,wsglob)


            #print("position: ",position)
            #print("fam: ",fam)
            #print("total_item_array[fam]: ",total_item_array[fam])

            if fam == 0:
                position = 0
            else:
                position = position + total_item_array[fam] + 5

            if total_item != 0:

                for row in ws.rows:
                    for cell in row:
                        new_cell = wsglob.cell(row=cell.row + position, column=cell.column,
                                value= cell.value)
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = copy(cell.number_format)
                            new_cell.protection = copy(cell.protection)
                            new_cell.alignment = copy(cell.alignment)

                wsprops = wsglob.sheet_properties
                wsprops.tabColor = "93c47d"
                wsprops.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)
                wsprops.outlinePr.applyStyles = True



            # Auto-sizing columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter # Get the column name
                for cell in col:
                    try: # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 1) * 1.5
                ws.column_dimensions[column].width = adjusted_width
                #print("Cell width adjusted for : ", cell)

        for col in wsglob.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 1) * 1.5
            wsglob.column_dimensions[column].width = adjusted_width
            #print("Cell width adjusted for : ", cell)

        print(total_item_array)

        output_filename =(path.replace(".xls", "-par famille.xls"))
        wb.save(output_filename)

        print("G??n??ration du fichier client termin??e :", output_filename)
        showerror(title='Error', message=("G??n??ration du fichier CLIENTS termin??e: " + output_filename))

root = App()
root.mainloop()
