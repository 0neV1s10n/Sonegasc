# Python program to read an excel file

# import openpyxl module
from queue import Empty
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# create a workbook
wb = Workbook()
#ws = wb.active

# Give the location of the file
path = "/Users/lqa/Downloads/Commande 29 octobre 2022.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path, data_only=True)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# Getting the value of maximum rows
# and column
row = sheet_obj.max_row
column = sheet_obj.max_column
  
print("Total Rows:", row)
print("Total Columns:", column)


# initialize family and family_pos_row arrays
family = []
family_pos_col = []

print("\n - FAMILIES with orders")
for i in range(1, column + 1): 
    cell_obj = sheet_obj.cell(row = 5, column = i) 
    orders = sheet_obj.cell(row = row, column = i) 
    if cell_obj.value and orders.value != 0:
        print(cell_obj.value, end = " ")
        family.append(cell_obj.value)
        family_pos_col.append(i)    

# printing the value of last row
print("\nValue of last column - PRODUCTORS LIST")

# initialize prod_name and prod_pos_row arrays
prod_name = []
prod_pos_row =[]

# loop through rows and detect productors and position in file
for a in range(1, row + 1):

    producteur = sheet_obj.cell(row = a, column = 1).value
    if producteur:
        if '"**"' in str(producteur):
            producteur=(producteur.lstrip('"**"'))
            prod_name.append(producteur)
            prod_pos_row.append(a)


# Loop through file to list ordered items per family
for fam in range (0,len(family)):
    total = 0

    print("*** Famille: ", family[fam]," ***")
    print("*** Colonne nr: ", family_pos_col[fam]," ***")
    wb.create_sheet(family[fam])
    if fam == 1: del wb['Sheet']

    ws = wb[family[fam]]

    ws['A1'] = 'INTITULE'
    ws['B1'] = 'DESCRIPTION'
    ws['C1'] = 'UNITE'
    ws['D1'] = 'PRIX UNITAIRE'
    ws['E1'] = 'QUANTITE'
    ws['F1'] = 'TOTAL'
        
    for prod in range (0,len(prod_name)-1):
        
        print ("Nom du producteur: ",prod_name[prod])
        print ("Première ligne du producteur: ",prod_pos_row[prod]+1)
        print ("Première ligne du producteur suivant: ",prod_pos_row[prod+1])

        ws.append((prod_name[prod],""))

        #Defining printing area
        ws_row = sheet_obj.max_row
        ws_column = sheet_obj.max_column

        last_cell=(get_column_letter(ws_column))
        print("A1:",last_cell,ws_column)
        ws.print_area = "A1:",(last_cell)

        for item in range (prod_pos_row[prod]+1,(prod_pos_row[prod+1])):

            intitulé = sheet_obj.cell(row = item, column = 1).value
            description = sheet_obj.cell(row = item, column = 2).value
            unite = sheet_obj.cell(row = item, column = 3).value
            prixun = sheet_obj.cell(row = item, column = 4).value
            quantite_total = sheet_obj.cell(row = item, column = family_pos_col[fam]).value
               
            if quantite_total and float(quantite_total) > 0:

                montant_total = float(quantite_total) * float(prixun)
                montant_total=round(montant_total,2)

                print("\nITEM Ligne: ",item)
                print("ITEM Colonne: ",family_pos_col[fam])
                print("ITEM Famille: ",family[fam])
                print("ITEM quantite_total: ",quantite_total)
                print("ITEM prix unitaire: ",prixun)
                print("montant_total: ",montant_total)

                print(intitulé,";",description,";",unite,";",prixun,";",quantite_total,";",montant_total)

                insert_row=intitulé,description,unite,prixun,quantite_total,montant_total
                ws.append(insert_row)

                total = total + montant_total
                
                # Auto-sizing columns
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter # Get the column name
                    for cell in col:
                        try: # Necessary to avoid error on empty cells
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[column].width = adjusted_width
                


    ligne_total="TOTAL",family[fam],"IBAN","BE33000441432246","",total
    ws.append(ligne_total)

wb.save('/Users/lqa/Downloads/familles.xlsx')