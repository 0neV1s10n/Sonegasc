# Python program to read an excel file

# import openpyxl module
import itertools
from operator import concat
from queue import Empty
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import Font, Color, colors, PatternFill, Border, Side, Alignment, Protection
from openpyxl.worksheet.table import Table, TableStyleInfo

# create a workbook
wb = Workbook()
#ws = wb.active

# Give the location of the file
path = "/Users/lqa/Downloads/Commande 29 octobre 2022.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path, data_only=True)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# Getting the value of maximum rows
# and column
row = sheet_obj.max_row
column = sheet_obj.max_column
  
print("Total Rows:", row)
print("Total Columns:", column)


print("\nValue of first row - FAMILY NAME")
for i in range(1, column + 1): 
    cell_obj = sheet_obj.cell(row = 5, column = i) 
    if cell_obj.value:
        print(cell_obj.value, end = " ")
    

# printing the value of last row
print("\nValue of last column - PRODUCTORS LIST")

# initialize prod_name and prod_pos_row arrays
prod_name = []
prod_pos_row =[]
sheets_to_remove = []


# loop through rows and detect productors and position in file
for a in range(1, row + 1):

    producteur = sheet_obj.cell(row = a, column = 1).value
    if producteur:
        if '"**"' in str(producteur):
            producteur=(producteur.lstrip('"**"'))
            prod_name.append(producteur)
            prod_pos_row.append(a)

# List productors and positions in file
#print(prod_name)
#print(prod_pos_row)

# Print headers / columns titles
#print("intitulé;description;unite;prix unité;quantite_total;montant_total")


# Loop through file to list ordered items per productor
for prod in range (0,len(prod_name)-1):

    total = 0
#    print("Producteur ",prod,"/",len(prod_name))
    print("*** Producteur: ", prod_name[prod]," ***")
    wb.create_sheet(prod_name[prod])
    if prod == 1: del wb['Sheet']



    ws = wb[prod_name[prod]]
    ws['A1'] = 'INTITULE'
    ws['B1'] = 'DESCRIPTION'
    ws['C1'] = 'UNITE'
    ws['D1'] = 'PRIX UNITAIRE'
    ws['E1'] = 'QUANTITE'
    ws['F1'] = 'TOTAL'

    for item in range (prod_pos_row[prod]+1,(prod_pos_row[prod+1])):

            intitulé = sheet_obj.cell(row = item, column = 1)
            description = sheet_obj.cell(row = item, column = 2)
            unite = sheet_obj.cell(row = item, column = 3)
            prixun = sheet_obj.cell(row = item, column = 4)
            montant_total = sheet_obj.cell(row = item, column = sheet_obj.max_column - 1)
            quantite_total = sheet_obj.cell(row = item, column = sheet_obj.max_column - 2)

            if montant_total.value: 

                #print (sheet_obj.cell(row = item, column = sheet_obj.max_column - 1).value)
                if montant_total.value: 
                        #print ("intitulé : ",intitulé.value)
                        #print ("description : ",description.value)
                        #print ("unite : ",unite.value)
                        #print ("prix unité : ",prixun.value)
                        #print ("quantite_total : ",quantite_total.value)
                        #print ("montant_total : ",montant_total.value)

                        print(intitulé.value,";",description.value,";",unite.value,";",prixun.value,";",quantite_total.value,";",montant_total.value)

                        insert_row=intitulé.value,description.value,unite.value,prixun.value,quantite_total.value,montant_total.value
                        ws.append(insert_row)

                        total = total + float(montant_total.value)
                        
                        
            if item == (prod_pos_row[prod+1]-1) and total == 0:
                print("!!!!!!!!!!!!!!!!!!!!!!!!!",prod_name[prod],item,prod_pos_row[prod])
                print("TOTAL: ", total)
                sheets_to_remove.append(prod_name[prod])

    ligne_total="TOTAL " + prod_name[prod],"","","","",total
    ws.append(ligne_total)

    # Merge last row for total display
    current_row = ws.max_row
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)

    # Highlight last row / total
    current_row = str(current_row) + ":" + str(current_row)
    for cell in ws[current_row]:
        print(cell)
        cell.font = Font(color="0e1643", bold=True, size=14)

    # Auto-sizing columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
    #    adjusted_width = (max_length + 2) * 1.2
        adjusted_width = (max_length + 1) * 1.5
        ws.column_dimensions[column].width = adjusted_width
        print("Cell width adjusted for : ", cell)

    # Setting print area
    def convertTuple(tup):
        str = ''.join(tup)
        return str

    last = ws.calculate_dimension()
    print(last)
    ws.print_area = last


    wsprops = ws.sheet_properties
    print(wsprops)
    wsprops.tabColor = "1072BA"
    wsprops.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)
    wsprops.pageSetUpPr.autoPageBreaks = True    
    wsprops.print_title_rows='1:1'
    print(wsprops)

    ws.sheet_view.showGridLines = True
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    for cell in ws["1:1"]:
        print(cell)
        cell.font = Font(color="0e1643", bold=True, size=14)
#       cell.fill = PatternFill(bgColor="00FFFF00", fill_type = "solid")

print ("Feuilles présentes dans le fichier: ", wb.sheetnames)
print("Feuilles vides à supprimer: ", sheets_to_remove)
print("Nombre de feuilles vides à supprimer: ", len(sheets_to_remove))

for rem in range(0,len(sheets_to_remove)):
    print(rem, sheets_to_remove[rem])
    wb.remove(wb[sheets_to_remove[rem]])

wb.save('/Users/lqa/Downloads/producteurs.xlsx')